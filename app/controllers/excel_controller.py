import os
from flask import jsonify, send_file
from flask_jwt_extended import jwt_required
from urllib.parse import quote
from app import app
from app.exceptions.exceptions import CustomAPIException
from app.models.models  import Project  # 假设 Project 模型在 models 模块中
from app.controllers.project_material_controller import get_project_materials_info
import openpyxl
from openpyxl.styles import Border, Side, Alignment
# 模板路径
TEMPLATE_PATH = os.path.join(app.config['TEMPLATE_FOLDER'], "materialtabletemplate.xlsx")
from datetime import datetime





@jwt_required()
def generate_excel(project_id):
    try:
        # 从数据库获取项目数据
        project = Project.query.get(project_id)
        if not project:
            return jsonify({"error": "项目不存在"}), 404

        # 获取项目关联的材料列表
        material_list = get_project_materials_info(project_id)
        if not material_list:
            return jsonify({"error": "材料列表为空"}), 400

        # 提取项目信息
        product_info = {
            "产品名称": project.project_name,
            "产品型号": project.project_model,
            "成品规格": project.project_type,
            "文件编号": project.file_number,
            "产品等级": "N/A",
            "产品编号": project.product_number,
        }
        today = datetime.today()

        # 生成格式化的日期字符串
        formatted_date = today.strftime("%y%m%d")

        print(formatted_date)
        # 动态生成输出文件名称
        product_model = project.project_model or "未命名产品"
        output_file_name = f"{product_model}BOM明细表 {formatted_date}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_file_name)

        # 调用填充函数
        fill_excel_template(TEMPLATE_PATH, output_path, product_info, material_list)

        # URL 编码文件名（避免中文乱码）
        encoded_file_name = quote(output_file_name)

        # 返回文件
        response = send_file(
            output_path,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # 设置 Content-Disposition 响应头
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_file_name}"
        return response
    except Exception as e:
        raise CustomAPIException("Material not found in the project", 404)




def fill_excel_template(template_path, output_path, product_info, data):
    """
    填充 Excel 模板数据，并将结果保存到指定路径。
    """
    # **加载 Excel 模板**
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    # **填充固定项目信息**
    def set_cell_value(cell, value):
        if not isinstance(sheet[cell], openpyxl.cell.cell.MergedCell):
            sheet[cell].value = value

    set_cell_value("C2", product_info.get("产品名称", ""))
    set_cell_value("C3", product_info.get("产品型号", ""))
    set_cell_value("C4", product_info.get("成品规格", ""))
    set_cell_value("H1", product_info.get("文件编号", ""))
    set_cell_value("H3", product_info.get("产品等级", ""))
    set_cell_value("H4", product_info.get("产品编号", ""))

    # **动态插入数据行**
    start_row = 6  # 物料数据从第6行开始
    required_rows = len(data) - 1  # 计算要插入的行数

    if required_rows > 0:
        sheet.insert_rows(start_row, amount=required_rows)  # **一次性插入 required_rows 行**

    # **手动合并 F:G、H:J**
    for idx in range(len(data)):
        current_row = start_row + idx
        sheet.merge_cells(f"F{current_row}:G{current_row}")  # **合并 F:G**
        sheet.merge_cells(f"H{current_row}:J{current_row}")  # **合并 H:J**

    # **填充数据**
    for idx, row_data in enumerate(data):
        current_row = start_row + idx
        sheet[f"A{current_row}"].value = idx + 1  # **自动生成序号**
        sheet[f"B{current_row}"].value = row_data.get("material_code", "")
        sheet[f"C{current_row}"].value = row_data.get("material_name", "")
        sheet[f"D{current_row}"].value = row_data.get("model_specification", "")
        sheet[f"E{current_row}"].value = row_data.get("unit", "")

    # **手动定义边框**
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + required_rows, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")  # **居中对齐**

    # **保存 Excel 文件**
    wb.save(output_path)
    print(f"文件已保存到: {output_path}")



def generate_excel_local(project_id):
    try:
        # 从数据库获取项目数据
        project = Project.query.get(project_id)
        if not project:
            return jsonify({"error": "项目不存在"}), 404

        # 获取项目关联的材料列表
        material_list = get_project_materials_info(project_id)
        if not material_list:
            return jsonify({"error": "材料列表为空"}), 400

        # 提取项目信息
        product_info = {
            "产品名称": project.project_name,
            "产品型号": project.project_model,
            "成品规格": project.project_type,
            "文件编号": project.file_number,
            "产品等级": project.project_level,
            "产品编号": project.product_number,
        }
        # 获取当前日期
        today = datetime.today()

        # 生成格式化的日期字符串
        formatted_date = today.strftime("%y%m%d")

        print(formatted_date)
        # 动态生成输出文件名称
        product_model = project.project_model or "未命名产品"
        output_file_name = f"{product_model}BOM明细表 {formatted_date}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_file_name)

        # 调用填充函数
        fill_excel_template(TEMPLATE_PATH, output_path, product_info, material_list)
        return output_path
    except Exception as e:
        raise CustomAPIException("Material not found in the project", 404)